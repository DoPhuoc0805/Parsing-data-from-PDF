"""
Đọc file Excel thành danh sách dòng thô (tầng bronze).

Đầu ra có cùng dạng với read_pdf.py — danh sách dict theo tên trường chuẩn —
nên toàn bộ phần xử lý phía sau (dựng cây phân cấp, kế thừa cột, gộp nhóm)
dùng lại y nguyên, không cần code riêng cho Excel.

Ba khác biệt so với PDF cần xử lý ở đây:

1. **Bảng không bắt đầu ở dòng đầu sheet.** Văn bản thường có tiêu đề, ghi
   chú phía trên; dòng header được tìm bằng chính tiêu chí nhận diện bảng
   nhiệm vụ ở fields.py (khớp đủ đơn vị chủ trì + đơn vị phối hợp).

2. **Ô gộp (merged cell).** openpyxl chỉ trả giá trị ở ô góc trên-trái, các ô
   còn lại trong vùng gộp là None. Giá trị được điền ra toàn vùng để dòng con
   nhìn thấy giá trị dùng chung — giống cách pdfplumber vốn đã trả về.

3. **Ô không phải chuỗi.** Excel lưu số và ngày tháng theo kiểu riêng, trong
   khi tầng bronze là JSON và các bước sau đều làm việc trên chuỗi. Số giữ
   nguyên cách hiển thị (vd 1.0 -> "1.0", vì phần thập phân chính là cấp trong
   mã phân cấp); ngày tháng đổi sang dd/mm/yyyy cho khớp với các ô cùng cột
   vốn được gõ tay theo định dạng đó.
"""

from __future__ import annotations

from datetime import date, datetime

import openpyxl

from .fields import is_task_header, match_header_fields

# Bảng có thể nằm sau vài dòng tiêu đề/ghi chú ở đầu sheet.
HEADER_SCAN_LIMIT = 10


def _cell_to_text(value):
    """Đưa mọi kiểu ô Excel về chuỗi (hoặc None) để tầng bronze là JSON hợp lệ."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    text = str(value).strip()
    return text or None


def _read_grid(worksheet) -> list:
    """Đọc sheet thành lưới 2 chiều, điền giá trị ô gộp ra toàn vùng gộp."""
    grid = [[cell.value for cell in row] for row in worksheet.iter_rows()]

    for merged in worksheet.merged_cells.ranges:
        value = grid[merged.min_row - 1][merged.min_col - 1]
        for row_index in range(merged.min_row - 1, merged.max_row):
            for col_index in range(merged.min_col - 1, merged.max_col):
                grid[row_index][col_index] = value

    return [[_cell_to_text(value) for value in row] for row in grid]


def _find_header_row(grid: list):
    """Trả về (chỉ số dòng header, column_mapping) hoặc (None, None)."""
    for row_index, row in enumerate(grid[:HEADER_SCAN_LIMIT]):
        column_mapping = match_header_fields([row])
        if is_task_header(column_mapping):
            return row_index, column_mapping
    return None, None


def read_excel(xlsx_path: str, sheet_name: str = None) -> list:
    """Đọc bảng nhiệm vụ trong file Excel ra danh sách dòng thô."""
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]

    grid = _read_grid(worksheet)
    header_index, column_mapping = _find_header_row(grid)
    if header_index is None:
        return []

    rows = []
    for row in grid[header_index + 1 :]:
        record = {
            field: (row[col_index] if col_index < len(row) else None)
            for col_index, field in column_mapping.items()
        }
        if any(value for value in record.values()):
            rows.append(record)

    return rows


if __name__ == "__main__":
    import json
    import sys

    # Console Windows mac dinh dung codepage khong ho tro tieng Viet.
    sys.stdout.reconfigure(encoding="utf-8")

    rows = read_excel(sys.argv[1])
    print(json.dumps(rows, ensure_ascii=False, indent=2))
    print(f"\n=> Tong so dong: {len(rows)}", file=sys.stderr)
